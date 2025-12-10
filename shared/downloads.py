import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font


def excel_format(data: pd.DataFrame, table_name: str) -> Workbook:
    # Create a workbook and add a new worksheet
    workbook = Workbook()
    sheet = workbook.active

    # Configure table header
    header_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(size=12, bold=True, color='00FFFFFF')

    # Write DataFrame headers to Excel sheet
    for col_idx, column in enumerate(data.columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=column)
        cell.alignment = header_align
        cell.font = header_font

    # Write DataFrame data to Excel sheet
    for r_idx, row in enumerate(data.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    # Set the column width to auto-size
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells) + 4
        sheet.column_dimensions[column_cells[0].column_letter].width = length

    # Convert the data range into an Excel table
    tab = Table(displayName=table_name, ref=sheet.dimensions)
    style = TableStyleInfo(name="TableStyleMedium16", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    sheet.add_table(tab)
    return workbook
